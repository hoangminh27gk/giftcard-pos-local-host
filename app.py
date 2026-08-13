"""
Gift Card POS - production web version
- PostgreSQL in the cloud (via DATABASE_URL), SQLite fallback for local use
- Username + password login (APP_USERNAME / APP_PASSWORD env vars)
- Deployable to Render/Railway with gunicorn

Local run:  pip install -r requirements.txt && python app.py
"""

import os
import io
import json
from datetime import datetime
from functools import wraps

from flask import (
    Flask, request, jsonify, render_template, session,
    redirect, url_for, send_file,
)
from sqlalchemy import (
    create_engine, MetaData, Table, Column, Integer, String, Float, text,
)

# ── Config ────────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATABASE_URL = os.environ.get("DATABASE_URL", "")
if DATABASE_URL.startswith("postgres://"):
    # Render provides postgres:// but SQLAlchemy needs postgresql://
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
if not DATABASE_URL:
    DATABASE_URL = "sqlite:///" + os.path.join(BASE_DIR, "giftcards.db")

APP_USERNAME = os.environ.get("APP_USERNAME", "admin")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "password123!")
SECRET_KEY = os.environ.get("SECRET_KEY", "dev-secret-change-me")

BUSINESS_NAME = os.environ.get("BUSINESS_NAME", "Giftcard POS")
BUSINESS_ADDRESS = os.environ.get("BUSINESS_ADDRESS", "")
BUSINESS_PHONE = os.environ.get("BUSINESS_PHONE", "")

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB uploads

engine = create_engine(DATABASE_URL, pool_pre_ping=True, future=True)
metadata = MetaData()

gift_cards = Table(
    "gift_cards", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("gc_code", String(64), unique=True, nullable=False),
    Column("msr_id", String(64)),
    Column("cust_mobile", String(32)),
    Column("balance", Float, nullable=False, default=0),
    Column("status", String(16), nullable=False, default="ACTIVE"),
    Column("created_at", String(32)),
)

transactions = Table(
    "transactions", metadata,
    Column("id", Integer, primary_key=True, autoincrement=True),
    Column("gc_code", String(64), nullable=False),
    Column("type", String(16), nullable=False),  # CHARGE | REFILL | IMPORT | REMOVE
    Column("amount", Float, nullable=False),
    Column("balance_before", Float),
    Column("balance_after", Float),
    Column("ticket_no", String(64)),
    Column("note", String(255)),
    Column("created_at", String(32)),
)

# Simple key/value store for settings editable from the Settings tab
# (business info, login credentials, printing preferences). Seeded once
# from the env vars above, then the DB is the source of truth.
app_settings = Table(
    "app_settings", metadata,
    Column("key", String(64), primary_key=True),
    Column("value", String(500)),
)

SETTINGS_DEFAULTS = {
    "business_name": BUSINESS_NAME,
    "business_address": BUSINESS_ADDRESS,
    "business_phone": BUSINESS_PHONE,
    "app_username": APP_USERNAME,
    "app_password": APP_PASSWORD,
    "print_mode": "browser",   # "browser" or "usb"
    "paper_width": "80",       # 58 or 80 (mm)
    "auto_print": "false",
}


def get_setting(key, default=""):
    with engine.connect() as conn:
        row = conn.execute(text("SELECT value FROM app_settings WHERE key=:k"), {"k": key}).fetchone()
    return row[0] if row else default


def set_setting(key, value):
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO app_settings (key, value) VALUES (:k, :v)
            ON CONFLICT (key) DO UPDATE SET value = excluded.value
        """), {"k": key, "v": value})


def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def init_db():
    metadata.create_all(engine)

    # Seed settings once from env-var defaults; DB is the source of truth after that
    with engine.begin() as conn:
        existing = {row[0] for row in conn.execute(text("SELECT key FROM app_settings"))}
        for key, value in SETTINGS_DEFAULTS.items():
            if key not in existing:
                conn.execute(app_settings.insert().values(key=key, value=value))

    # Seed once from seed_data.json if the database is empty
    seed_path = os.path.join(BASE_DIR, "seed_data.json")
    if not os.path.exists(seed_path):
        return
    with engine.begin() as conn:
        count = conn.execute(text("SELECT COUNT(*) FROM gift_cards")).scalar()
        if count and count > 0:
            return
        with open(seed_path, encoding="utf-8") as f:
            seed = json.load(f)
        for c in seed.get("gift_cards", []):
            conn.execute(gift_cards.insert().values(
                gc_code=str(c["gc_code"]),
                msr_id=c.get("msr_id") or "",
                cust_mobile=c.get("cust_mobile") or "",
                balance=float(c.get("balance") or 0),
                status=c.get("status") or "ACTIVE",
                created_at=c.get("created_at") or now_str(),
            ))
        for t in seed.get("transactions", []):
            conn.execute(transactions.insert().values(
                gc_code=str(t["gc_code"]),
                type=t.get("type") or "IMPORT",
                amount=float(t.get("amount") or 0),
                balance_before=t.get("balance_before"),
                balance_after=t.get("balance_after"),
                ticket_no=t.get("ticket_no") or "",
                note=t.get("note") or "",
                created_at=t.get("created_at") or now_str(),
            ))
        print(f"Seeded {len(seed.get('gift_cards', []))} cards from seed_data.json")


init_db()


# ── Auth ──────────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "Not logged in.", "auth": False}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if username == get_setting("app_username", APP_USERNAME) and password == get_setting("app_password", APP_PASSWORD):
            session["logged_in"] = True
            session["username"] = username
            session.permanent = True
            return redirect(url_for("index"))
        return render_template("login.html", error="Incorrect username or password. Please try again.")
    if session.get("logged_in"):
        return redirect(url_for("index"))
    return render_template("login.html", error=None)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── Helpers ───────────────────────────────────────────────────────────────────

def row_to_dict(row):
    return dict(row._mapping)


def normalize_gc_code(code):
    """GC codes are 00001-99999. Pad short numeric entries ("1", "01", "001")
    up to 5 digits so "1" and "00001" always refer to the same card."""
    code = (code or "").strip()
    if code.isdigit() and len(code) <= 5:
        return code.zfill(5)
    return code


def get_card(conn, code):
    return conn.execute(
        text("SELECT * FROM gift_cards WHERE gc_code = :c"), {"c": code}
    ).fetchone()


# ── Pages ─────────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    return render_template(
        "index.html",
        business_name=get_setting("business_name", BUSINESS_NAME),
        business_address=get_setting("business_address", BUSINESS_ADDRESS),
        business_phone=get_setting("business_phone", BUSINESS_PHONE),
    )


@app.route("/healthz")
def healthz():
    return "ok", 200


# ── API ───────────────────────────────────────────────────────────────────────

@app.route("/api/lookup", methods=["POST"])
@login_required
def lookup():
    data = request.json or {}
    code = (data.get("gc_code") or "").strip()
    if not code:
        return jsonify({"error": "Please enter a gift card code."}), 400

    with engine.connect() as conn:
        card = get_card(conn, normalize_gc_code(code))

        # Not a GC code? Try MSR ID — full value or just the last digits (e.g. "62")
        if not card:
            rows = conn.execute(
                text("SELECT * FROM gift_cards WHERE msr_id = :m OR msr_id LIKE :pat ORDER BY msr_id LIMIT 20"),
                {"m": code, "pat": f"%{code}"},
            ).fetchall()
            if len(rows) == 1:
                card = rows[0]
            elif len(rows) > 1:
                return jsonify({"matches": [
                    {"gc_code": r._mapping["gc_code"], "msr_id": r._mapping["msr_id"],
                     "balance": r._mapping["balance"], "status": r._mapping["status"]}
                    for r in rows
                ]})

        if not card:
            return jsonify({"error": f"No card found with code or MSR ID '{code}'."}), 404

        gc_code = card._mapping["gc_code"]
        txns = conn.execute(
            text("SELECT * FROM transactions WHERE gc_code = :c ORDER BY id DESC LIMIT 10"),
            {"c": gc_code},
        ).fetchall()

    return jsonify({
        "card": row_to_dict(card),
        "transactions": [row_to_dict(t) for t in txns],
    })


@app.route("/api/charge", methods=["POST"])
@login_required
def charge():
    data = request.json or {}
    code = normalize_gc_code(data.get("gc_code"))
    amount = data.get("amount")
    ticket_no = (data.get("ticket_no") or "").strip()
    note = (data.get("note") or "").strip()

    if not code:
        return jsonify({"error": "Gift card code is required."}), 400
    try:
        amount = float(amount)
        if amount <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"error": "Amount must be a positive number."}), 400

    with engine.begin() as conn:
        card = get_card(conn, code)
        if not card:
            return jsonify({"error": f"Card '{code}' not found."}), 404
        card = row_to_dict(card)
        if card["status"] == "USED":
            return jsonify({"error": "This card has already been fully used."}), 400
        if card["balance"] <= 0:
            return jsonify({"error": "This card has a $0.00 balance."}), 400
        if amount > card["balance"] + 1e-9:
            return jsonify({"error": f"Charge ${amount:.2f} exceeds balance ${card['balance']:.2f}."}), 400

        new_balance = round(card["balance"] - amount, 2)
        new_status = "USED" if new_balance == 0 else "PARTIALUSED"

        conn.execute(
            text("UPDATE gift_cards SET balance=:b, status=:s WHERE gc_code=:c"),
            {"b": new_balance, "s": new_status, "c": code},
        )
        conn.execute(transactions.insert().values(
            gc_code=code, type="CHARGE", amount=amount,
            balance_before=card["balance"], balance_after=new_balance,
            ticket_no=ticket_no, note=note, created_at=now_str(),
        ))

    return jsonify({
        "success": True, "gc_code": code, "charged": amount,
        "balance_before": card["balance"], "balance_after": new_balance,
        "status": new_status, "ticket_no": ticket_no,
    })


@app.route("/api/refill", methods=["POST"])
@login_required
def refill():
    data = request.json or {}
    code = normalize_gc_code(data.get("gc_code"))
    amount = data.get("amount")
    note = (data.get("note") or "").strip()

    if not code:
        return jsonify({"error": "Gift card code is required."}), 400
    try:
        amount = float(amount)
        if amount <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"error": "Refill amount must be a positive number."}), 400

    with engine.begin() as conn:
        card = get_card(conn, code)
        if not card:
            return jsonify({"error": f"Card '{code}' not found."}), 404
        card = row_to_dict(card)

        new_balance = round(card["balance"] + amount, 2)
        conn.execute(
            text("UPDATE gift_cards SET balance=:b, status='ACTIVE' WHERE gc_code=:c"),
            {"b": new_balance, "c": code},
        )
        conn.execute(transactions.insert().values(
            gc_code=code, type="REFILL", amount=amount,
            balance_before=card["balance"], balance_after=new_balance,
            ticket_no="", note=note, created_at=now_str(),
        ))

    return jsonify({
        "success": True, "gc_code": code, "refilled": amount,
        "balance_before": card["balance"], "balance_after": new_balance,
        "status": "ACTIVE",
    })


@app.route("/api/add-card", methods=["POST"])
@login_required
def add_card():
    data = request.json or {}
    code = normalize_gc_code(data.get("gc_code"))
    msr_id = (data.get("msr_id") or "").strip()
    mobile = (data.get("cust_mobile") or "").strip()
    balance = data.get("balance", 0)
    status = (data.get("status") or "ACTIVE").strip()

    if not code:
        return jsonify({"error": "GC Code is required."}), 400
    try:
        balance = float(balance)
        if balance < 0:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"error": "Balance must be 0 or a positive number."}), 400

    if status not in ("ACTIVE", "PARTIALUSED", "USED"):
        status = "ACTIVE"

    with engine.begin() as conn:
        existing = get_card(conn, code)
        if existing:
            return jsonify({"error": f"Card '{code}' already exists in the database."}), 409
        conn.execute(gift_cards.insert().values(
            gc_code=code, msr_id=msr_id, cust_mobile=mobile,
            balance=balance, status=status, created_at=now_str(),
        ))
        conn.execute(transactions.insert().values(
            gc_code=code, type="IMPORT", amount=balance,
            balance_before=0, balance_after=balance,
            ticket_no="", note="Manually added", created_at=now_str(),
        ))

    return jsonify({"success": True, "gc_code": code, "balance": balance, "status": status})


@app.route("/api/update-card", methods=["POST"])
@login_required
def update_card():
    data = request.json or {}
    code = normalize_gc_code(data.get("gc_code"))
    if not code:
        return jsonify({"error": "Gift card code is required."}), 400

    with engine.begin() as conn:
        card = get_card(conn, code)
        if not card:
            return jsonify({"error": f"Card '{code}' not found."}), 404
        card = row_to_dict(card)

        # Only update the fields that were provided
        msr_id = data.get("msr_id")
        mobile = data.get("cust_mobile")
        new_msr = msr_id.strip() if isinstance(msr_id, str) else card["msr_id"]
        new_mobile = mobile.strip() if isinstance(mobile, str) else card["cust_mobile"]

        conn.execute(
            text("UPDATE gift_cards SET msr_id=:m, cust_mobile=:mo WHERE gc_code=:c"),
            {"m": new_msr, "mo": new_mobile, "c": code},
        )

    return jsonify({"success": True, "gc_code": code,
                    "msr_id": new_msr, "cust_mobile": new_mobile})


@app.route("/api/remove-card", methods=["POST"])
@login_required
def remove_card():
    data = request.json or {}
    code = normalize_gc_code(data.get("gc_code"))
    if not code:
        return jsonify({"error": "Gift card code is required."}), 400

    with engine.begin() as conn:
        card = get_card(conn, code)
        if not card:
            return jsonify({"error": f"Card '{code}' not found."}), 404
        card = row_to_dict(card)
        conn.execute(text("DELETE FROM gift_cards WHERE gc_code = :c"), {"c": code})
        # Keep an audit record of the removal
        conn.execute(transactions.insert().values(
            gc_code=code, type="REMOVE", amount=card["balance"],
            balance_before=card["balance"], balance_after=0,
            ticket_no="", note="Card removed", created_at=now_str(),
        ))

    return jsonify({"success": True, "gc_code": code, "removed_balance": card["balance"]})


@app.route("/api/import-template")
@login_required
def import_template():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed on the server."}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gift Cards"

    headers = ["GC Code", "MSR ID", "Cust Mobile", "Current Balance", "Status"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    # Example row to show the expected format — delete this row before importing
    ws.append(["00001", "1000-0001-05", "5551234567", 50.00, "ACTIVE"])

    widths = [18, 16, 16, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="giftcard_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# Column aliases the importer will recognize, so real-world spreadsheets that
# don't exactly match the downloadable template (different case, wording,
# extra spaces) still import instead of failing.
IMPORT_COLUMN_ALIASES = {
    "gc_code": ["gc code", "giftcard code", "gift card code", "card code", "code"],
    "msr_id": ["msr id", "msr", "card id", "swipe id"],
    "mobile": ["cust mobile", "customer mobile", "mobile", "phone", "phone number", "cust phone"],
    "balance": ["current balance", "balance", "amount", "current amount"],
    "status": ["status", "card status"],
}


def _normalize_header(v):
    return " ".join(str(v).strip().lower().split()) if v is not None else ""


def _find_import_columns(headers):
    normalized = [_normalize_header(h) for h in headers]
    found = {}
    for field, aliases in IMPORT_COLUMN_ALIASES.items():
        idx = next((normalized.index(a) for a in aliases if a in normalized), None)
        found[field] = idx
    return found


@app.route("/api/import-excel", methods=["POST"])
@login_required
def import_excel():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded."}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload an .xlsx or .xls file."}), 400

    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl not installed on the server."}), 500

    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return jsonify({
            "error": "Could not read that file. If it's an older .xls file, "
                     "open it in Excel and use File → Save As → Excel Workbook (.xlsx), then upload that."
        }), 400

    header_row = None
    columns = None
    try:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not row:
                continue
            cols = _find_import_columns(row)
            if cols["gc_code"] is not None:
                header_row = i
                columns = cols
                break
    except Exception:
        return jsonify({"error": "This file appears to be empty or corrupted."}), 400

    if header_row is None:
        return jsonify({
            "error": "Could not find a header row with a gift card code column "
                     "(expected something like 'GC Code'). Check that the first row of "
                     "your sheet has column titles, or download the template for the exact format."
        }), 400

    ci_gc, ci_msr = columns["gc_code"], columns["msr_id"]
    ci_mobile, ci_bal, ci_status = columns["mobile"], columns["balance"], columns["status"]
    if ci_bal is None:
        return jsonify({"error": "Could not find a balance column (expected something like 'Current Balance')."}), 400

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    added, updated, skipped = 0, 0, 0
    with engine.begin() as conn:
        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
            if not row:
                continue
            try:
                raw_code = cell(row, ci_gc)
                if isinstance(raw_code, float) and raw_code.is_integer():
                    raw_code = int(raw_code)  # Excel often turns "00001" into the number 1
                gc_code = normalize_gc_code(str(raw_code)) if raw_code is not None else ""
                if not gc_code or gc_code.lower() in ("none", "nan"):
                    continue

                raw_bal = cell(row, ci_bal)
                try:
                    balance = float(raw_bal) if raw_bal is not None and str(raw_bal).strip() != "" else 0.0
                except (ValueError, TypeError):
                    balance = 0.0

                raw_msr = cell(row, ci_msr)
                raw_mobile = cell(row, ci_mobile)
                raw_status = cell(row, ci_status)
                msr_id = str(raw_msr).strip() if raw_msr not in (None, "") else ""
                mobile = str(raw_mobile).strip() if raw_mobile not in (None, "") else ""
                status = str(raw_status).strip().upper() if raw_status not in (None, "") else "ACTIVE"
                if status not in ("ACTIVE", "PARTIALUSED", "USED"):
                    status = "ACTIVE"

                existing = get_card(conn, gc_code)
                if existing:
                    conn.execute(
                        text("UPDATE gift_cards SET balance=:b, status=:s, msr_id=:m, cust_mobile=:mo WHERE gc_code=:c"),
                        {"b": balance, "s": status, "m": msr_id, "mo": mobile, "c": gc_code},
                    )
                    updated += 1
                else:
                    conn.execute(gift_cards.insert().values(
                        gc_code=gc_code, msr_id=msr_id, cust_mobile=mobile,
                        balance=balance, status=status, created_at=now_str(),
                    ))
                    conn.execute(transactions.insert().values(
                        gc_code=gc_code, type="IMPORT", amount=balance,
                        balance_before=0, balance_after=balance,
                        ticket_no="", note="Imported from Excel", created_at=now_str(),
                    ))
                    added += 1
            except Exception:
                skipped += 1
                continue

    message = f"Imported {added} new cards, updated {updated} existing."
    if skipped:
        message += f" Skipped {skipped} row(s) that couldn't be read."
    return jsonify({"success": True, "imported": added, "updated": updated, "skipped": skipped, "message": message})


@app.route("/api/all-cards")
@login_required
def all_cards():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()

    sql = "SELECT * FROM gift_cards WHERE 1=1"
    params = {}
    if q:
        sql += " AND (gc_code LIKE :q OR msr_id LIKE :q OR cust_mobile LIKE :q)"
        params["q"] = f"%{q}%"
    if status:
        sql += " AND status = :s"
        params["s"] = status
    sql += " ORDER BY id DESC LIMIT 500"

    with engine.connect() as conn:
        cards = conn.execute(text(sql), params).fetchall()
    return jsonify([row_to_dict(c) for c in cards])


@app.route("/api/settings", methods=["GET"])
@login_required
def get_settings():
    return jsonify({k: get_setting(k, v) for k, v in SETTINGS_DEFAULTS.items()})


@app.route("/api/settings", methods=["POST"])
@login_required
def update_settings():
    data = request.json or {}
    updated = {}
    for key in SETTINGS_DEFAULTS:
        if key in data:
            value = "" if data[key] is None else str(data[key])
            set_setting(key, value)
            updated[key] = value
    return jsonify({"success": True, **updated})


@app.route("/api/transactions", methods=["DELETE"])
@login_required
def clear_transactions():
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM transactions"))
    return jsonify({"success": True})


@app.route("/api/transactions")
@login_required
def all_transactions():
    q = request.args.get("q", "").strip()
    ttype = request.args.get("type", "").strip()

    sql = "SELECT * FROM transactions WHERE 1=1"
    params = {}
    if q:
        sql += " AND gc_code LIKE :q"
        params["q"] = f"%{q}%"
    if ttype:
        sql += " AND type = :t"
        params["t"] = ttype
    sql += " ORDER BY id DESC LIMIT 500"

    with engine.connect() as conn:
        txns = conn.execute(text(sql), params).fetchall()
    return jsonify([row_to_dict(t) for t in txns])


@app.route("/api/stats")
@login_required
def stats():
    with engine.connect() as conn:
        row = conn.execute(text("""
            SELECT
              COUNT(*) as total,
              SUM(CASE WHEN status='ACTIVE' THEN 1 ELSE 0 END) as active,
              SUM(CASE WHEN status='PARTIALUSED' THEN 1 ELSE 0 END) as partial,
              SUM(CASE WHEN status='USED' THEN 1 ELSE 0 END) as used,
              ROUND(CAST(SUM(balance) AS NUMERIC), 2) as total_balance
            FROM gift_cards
        """)).fetchone()
    d = row_to_dict(row)
    d["total_balance"] = float(d["total_balance"] or 0)
    return jsonify(d)


# -- Main (local dev) --

if __name__ == "__main__":
    print("\n" + "=" * 50)
    print("  Gift Card POS - Web Version")
    print("  Open your browser to: http://localhost:5000")
    print(f"  Login: {APP_USERNAME} / {APP_PASSWORD}")
    print("=" * 50 + "\n")
    app.run(debug=False, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
